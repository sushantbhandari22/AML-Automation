"""
core/styler.py — Professional Excel styling for AML reports.

Applies fonts, borders, number formatting, column widths, and freeze panes
to the final workbook. Completely decoupled from data logic.
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelStyler:
    """Apply professional banking-grade styling to an AML workbook."""

    # ── Colour Palette ───────────────────────────────────────────────
    HEADER_BG = "1F4E78"
    HEADER_FG = "FFFFFF"
    ALT_ROW = "F2F2F2"
    BORDER_CLR = "CCCCCC"
    AUDIT_BG = "333333"

    # ── Public Entry Point ───────────────────────────────────────────
    @classmethod
    def style_workbook(cls, file_path: str) -> None:
        """Read a saved workbook and apply styling to every sheet."""
        wb = load_workbook(file_path)
        border = cls._border()

        for name in wb.sheetnames:
            ws = wb[name]

            if name == "REPORT_SUMMARY":
                cls._style_summary(ws)
            elif name == "TreeMap":
                cls._style_tree_map(ws, border)
                ws.column_dimensions["B"].width = 5
                ws.column_dimensions["C"].width = 5
            elif name in ("MAIN_DATASET", "WORKING_DATASET", "PIVOT_YEAR", "PIVOT_CHANNEL", "PIVOT_MOBILE"):
                cls._style_audit(ws, border)
            else:
                cls._style_annex(ws, border)
                if name == "Annex-II":
                    cls._add_annex2_totals(ws, border)

            # Auto-adjust column widths (skip summary & tree map)
            if name not in ("REPORT_SUMMARY", "TreeMap"):
                cls._auto_width(ws)

        wb.save(file_path)

    # ── Private helpers ──────────────────────────────────────────────
    @classmethod
    def _border(cls):
        s = Side(style='thin', color=cls.BORDER_CLR)
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Summary Sheet ────────────────────────────────────────────────
    @classmethod
    def _style_summary(cls, ws):
        for c in (1, 2):
            ws.column_dimensions[get_column_letter(c)].width = 30
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=18, bold=True, color=cls.HEADER_BG)

    # ── Standard Annex (I-IV) ────────────────────────────────────────
    @classmethod
    def _style_annex(cls, ws, border):
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=16, bold=True, color=cls.HEADER_BG)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Metadata labels
        for r in range(2, 11):
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")

        # Data header (row 12)
        hdr = 12
        ws.row_dimensions[hdr].height = 28
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=hdr, column=c)
            cell.fill = PatternFill(start_color=cls.HEADER_BG, end_color=cls.HEADER_BG, fill_type="solid")
            cell.font = Font(bold=True, color=cls.HEADER_FG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Data rows
        for r in range(hdr + 1, ws.max_row + 1):
            alt = (r - hdr) % 2 == 0
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if alt:
                    cell.fill = PatternFill(start_color=cls.ALT_ROW, end_color=cls.ALT_ROW, fill_type="solid")

                hdr_val = str(ws.cell(row=hdr, column=c).value or "").lower()
                if any(k in hdr_val for k in ("debit", "credit", "balance", "sum", "amount", "total")):
                    cell.number_format = '#,##0.00;[Red]-#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        ws.freeze_panes = ws.cell(row=hdr + 1, column=1)

    # ── Annex II Totals Row ──────────────────────────────────────────
    @classmethod
    def _add_annex2_totals(cls, ws, border):
        last = ws.max_row
        data_start = 13
        ws.cell(row=last + 1, column=2).value = "TOTAL"
        ws.cell(row=last + 1, column=2).font = Font(bold=True)
        for ci in (3, 4, 5, 6):
            cl = get_column_letter(ci)
            cell = ws.cell(row=last + 1, column=ci)
            cell.value = f"=SUM({cl}{data_start}:{cl}{last})"
            cell.font = Font(bold=True)
            cell.border = border
            cell.number_format = '#,##0.00'

    # ── TreeMap ──────────────────────────────────────────────────────
    @classmethod
    def _style_tree_map(cls, ws, border):
        for r in range(1, 4):
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            cell = ws.cell(row=r, column=1)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=14, bold=True, color=cls.HEADER_BG) if r == 1 else Font(bold=True)

        for c in (1, 4):
            cell = ws.cell(row=5, column=c)
            cell.fill = PatternFill(start_color=cls.HEADER_BG, end_color=cls.HEADER_BG, fill_type="solid")
            cell.font = Font(bold=True, color=cls.HEADER_FG)
            cell.alignment = Alignment(horizontal="center")

        for r in range(7, ws.max_row + 1):
            for c in (1, 4):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif val:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.border = border

    # ── Audit Sheets ─────────────────────────────────────────────────
    @classmethod
    def _style_audit(cls, ws, border):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = PatternFill(start_color=cls.AUDIT_BG, end_color=cls.AUDIT_BG, fill_type="solid")
            cell.font = Font(bold=True, color=cls.HEADER_FG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.freeze_panes = "A2"

    # ── Auto Column Width ────────────────────────────────────────────
    @staticmethod
    def _auto_width(ws):
        for ci, col in enumerate(ws.columns, 1):
            mx = max((len(str(c.value)) for c in col if c.value), default=8)
            letter = get_column_letter(ci)
            hdr_val = str(ws.cell(row=12, column=ci).value or "").lower()
            cap = 50 if ("description" in hdr_val or "desc" in hdr_val) else 60
            ws.column_dimensions[letter].width = min(mx + 4, cap)
